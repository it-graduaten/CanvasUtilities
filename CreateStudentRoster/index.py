import os

import openpyxl
from canvasapi import Canvas


def ask_yes_no(question, default="y"):
    """ Asks the user a yes/no question """
    if default == "y":
        question = "{} [Y/n]: ".format(question)
    elif default == "n":
        question = "{} [y/N]: ".format(question)
    else:
        raise ValueError("Default value should be 'y' or 'n'")
    while True:
        choice = input(question).lower()
        if choice in ["y", "yes"]:
            return True
        elif choice in ["n", "no"]:
            return False
        elif choice == "":
            return default == "y"
        else:
            print("Please enter yes or no")


def ask_csv_or_excel():
    """ Asks the user if they need a csv or excel file """
    print("Do you need a csv or excel file?")
    types = ["csv", "excel"]
    for i, t in enumerate(types):
        print("{}. {}".format(i, t))
    while True:
        try:
            choice = int(input())
            if choice >= 0 and choice < len(types):
                return types[choice]
            else:
                print("Please enter a number between 0 and {}".format(
                    len(types) - 1))
        except ValueError:
            print("Please enter a number between 0 and {}".format(
                len(types) - 1))


def ask_horizontal_or_vertical():
    """ Asks the user if they need a horizontal or vertical file """
    print("Do you need a horizontal or vertical file?")
    types = ["horizontal", "vertical"]
    for i, t in enumerate(types):
        print("{}. {}".format(i, t))
    while True:
        try:
            choice = int(input())
            if choice >= 0 and choice < len(types):
                return types[choice]
            else:
                print("Please enter a number between 0 and {}".format(
                    len(types) - 1))
        except ValueError:
            print("Please enter a number between 0 and {}".format(
                len(types) - 1))


def create_csv_file(all_sections, course_id, include_section, include_student_number, include_student_name):
    # Create the csv file
    with open(f"studenten-{course_id}.csv", "w") as f:
        header = ""
        if include_section:
            header += "Klasgroep,"
        if include_student_number:
            header += "Studentennummer,"
        if include_student_name:
            header += "Naam,"
        # Remove last comma from header
        header = header[:-1]
        # Add the header
        f.write(header)
        # sort the students by their 'sortable_name'
        all_sections.sort(key=lambda s: s.name)
        for section in all_sections:
            section.students.sort(key=lambda s: s['sortable_name'])
        # Add a student to the csv file on every line
        for section in all_sections:
            for student in section.students:
                line = "\n"
                if include_section:
                    line += f"{section.name},"
                if include_student_number:
                    student_number = student['sis_user_id'].replace("S_", "")
                    line += f"{student_number},"
                if include_student_name:
                    line += f"{student['name']},"
                # Remove last comma from line
                line = line[:-1]
                # Add the line to the file
                f.write(line)
    f.close()


def ask_sections_to_include(all_sections):
    """ Asks the user which sections to include """
    all_sections = list(all_sections)
    sections_to_include = []
    for i, section in enumerate(all_sections):
        print("{}. {}".format(i, section.name))

    selected_sections = input("Which sections do you want to include (comma separated)? ")

    if selected_sections == "":
        return all_sections

    selected_sections = selected_sections.split(",")
    for section in selected_sections:
        try:
            sections_to_include.append(all_sections[int(section)])
        except ValueError:
            print("Please enter a number between 0 and {}".format(len(all_sections) - 1))
            exit(1)
    return sections_to_include


def create_excel_file(all_sections, course_id, include_section, include_student_number, include_student_name,
                      vertical_horizontal):
    # Create the Excel file
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Studenten"
    if vertical_horizontal == "vertical":
        row = 1
        col = 1
        # Add the header
        if include_section:
            sheet.cell(row=row, column=col).value = "Klasgroep"
            col += 1
        if include_student_number:
            sheet.cell(row=row, column=col).value = "Studentennummer"
            col += 1
        if include_student_name:
            sheet.cell(row=row, column=col).value = "Naam"
            col += 1
        sheet.cell(row=row, column=col).value = "Actief"
        # sort the students by their 'sortable_name'
        all_sections.sort(key=lambda s: s.name)
        for section in all_sections:
            section.students.sort(key=lambda s: s['sortable_name'])
        # Add the students to the excel file
        col = 1
        row = 2
        for section in all_sections:
            for student in section.students:
                if include_section:
                    sheet.cell(row=row, column=col).value = section.name
                    col += 1
                if include_student_number:
                    student_number = student['sis_user_id'].replace("S_", "")
                    sheet.cell(row=row, column=col).value = student_number
                    col += 1
                if include_student_name:
                    sheet.cell(row=row, column=col).value = student['name']
                    col += 1
                sheet.cell(row=row, column=col).value = "Ja" if student['enrollments'][0][
                                                                    'enrollment_state'] == "active" else "Nee"
                col = 1
                row += 1
    elif vertical_horizontal == "horizontal":
        row = 1
        col = 1
        # Add the header
        if include_section:
            sheet.cell(row=row, column=col).value = "Klasgroep"
            row += 1
        if include_student_number:
            sheet.cell(row=row, column=col).value = "Studentennummer"
            row += 1
        if include_student_name:
            sheet.cell(row=row, column=col).value = "Naam"
            row += 1
        sheet.cell(row=row, column=col).value = "Actief"
        # sort the students by their 'sortable_name'
        all_sections.sort(key=lambda s: s.name)
        for section in all_sections:
            section.students.sort(key=lambda s: s['sortable_name'])
        # Add the students to the excel file
        col = 2
        row = 1
        for section in all_sections:
            for student in section.students:
                if include_section:
                    sheet.cell(row=row, column=col).value = section.name
                    row += 1
                if include_student_number:
                    student_number = student['sis_user_id'].replace("S_", "")
                    sheet.cell(row=row, column=col).value = student_number
                    row += 1
                if include_student_name:
                    sheet.cell(row=row, column=col).value = student['name']
                    row += 1
                sheet.cell(row=row, column=col).value = "Ja" if student['enrollments'][0][
                                                                    'enrollment_state'] == "active" else "Nee"
                row = 1
                col += 1
    # Save the file
    wb.save(f"studenten-{course_id}.xlsx")


if __name__ == "__main__":
    # Get the canvas api key from the environment variables
    canvas_api_key = os.environ.get("CANVAS_API_KEY")
    # Login to canvas
    canvas = Canvas("https://thomasmore.instructure.com/", canvas_api_key)
    # Get the course
    canvas_course_id = input("Enter the Canvas course ID: ")
    course = canvas.get_course(canvas_course_id)
    # Get all students in the course
    sections = course.get_sections(include=['students', 'enrollments'])
    # Ask the user which sections to include
    sections = ask_sections_to_include(sections)
    # Ask the user if they want a csv or Excel file
    file_type = ask_csv_or_excel()
    include_sections = ask_yes_no("Do you want to include sections?")
    include_number = ask_yes_no("Do you want to include student numbers?")
    include_name = ask_yes_no("Do you want to include student names?")
    if file_type == "csv":
        create_csv_file(sections, canvas_course_id, include_sections, include_number, include_name)
    elif file_type == "excel":
        vertical_horizontal = ask_horizontal_or_vertical()
        create_excel_file(sections, canvas_course_id, include_sections, include_number, include_name,
                          vertical_horizontal)
    else:
        raise ValueError("File type should be 'csv' or 'excel'")
